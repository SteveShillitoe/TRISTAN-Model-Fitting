"""
This class module provides the functionality for the creation
of an Excel spreadsheet to store the results from the batch
processing of time-concentration data files. 

This is done using the openpyxl Python package.
"""

from openpyxl import Workbook
import logging

logger = logging.getLogger(__name__)

class ExcelWriter:
    def __init__(self, fullFilePath): 
        """Creates an instance of the ExcelWriter class that
       contains an empty Excel Workbook.
       
       Input Parameter
       ----------------
       fullFilePath - location where the Excel spreadsheet will be stored.
       """
        try:
            self.fullFilePath = fullFilePath
            #print ("Excel fullFilePath = " + fullFilePath)
            self.wb = Workbook()
            
            logger.info('In module ' + __name__ 
                    + '. Created an instance of class ExcelWriter.')
        except Exception as e:
            print('ExcelWriter.__init__: ' + str(e)) 
            logger.error('ExcelWriter.__init__: ' + str(e)) 

    def isWorksheet(self, title) -> bool:
        """Returns True if the worksheet called title already 
        exists in the Excel workbook."""
        boolWSExists = False
        for sheet in self.wb:
            if sheet.title == title:
                boolWSExists = True
                break
        
        return boolWSExists

    def RecordSkippedFiles(self, fileName, failureReason):
        """Records details of CSV data files that have to be skipped
       during batch processing because they do not conform
       to required format in a worksheet called 'Skipped files'.
       
       Input Parameters
       ----------------
       fileName - Name of the skipped file.
       failureReason - String containing the reason why the
      file was skipped."""
        try:
            if not self.isWorksheet("Skipped files"):
                #Skipped files tab does not exist.
                #Create it and add first skipped file info
                #to the first row on the worksheet.
                self.ws = self.wb.active
                self.ws.title = "Skipped files"
                self.ws['A1'] = "File"
                self.ws['B1'] = "Failure Reason"
                self.ws['A2'] = fileName
                self.ws['B2'] = failureReason
            else:
                thisWS = self.wb["Skipped files"]
                #get next empty row
                row_count = thisWS.max_row
                nextRow = row_count + 1
                thisWS['A' + str(nextRow)] = fileName
                thisWS['B' + str(nextRow)] = failureReason
                    
            logger.info('In module ' + __name__ 
                    + '.RecordSkippedFiles.')

        except Exception as e:
            print('ExcelWriter.RecordSkippedFiles: ' + str(e)) 
            logger.error('ExcelWriter.RecordSkippedFiles: ' + str(e)) 

    def RecordParameterValues(self, fileName, modelName, paramName, 
                              paramValue, paramLower, paramUpper):
        """During batch processing, records each optimum parameter 
        value (and associated information) resulting from curve 
        fitting in a row on a worksheet dedicated to that parameter. 
        Creates that worksheet when the parameter value from the
        first data file being processed is available. 
        
        Input Parameters
        -----------------
        fileName - Name of the data file currently being processed.
        modelName - Name of the model used for curve fitting.
        paramName - Name of the parameter.
        paramValue - Value of the parameter resulting from curve fitting.
        paramLower, paramUpper - Lower and upper 95% confidence interval
                of the paramValue. 
        """
        try:
            paramName = paramName.partition(',')
            paramName = paramName[0] 
            paramName = paramName.replace('\'', '')
            paramName = paramName[0:31] #worksheet title max 31 chars 
                
            if not self.isWorksheet(paramName):
                #This parameter tab does not exist.
                #Create it and add first parameter value data
                #to the first row of the worksheet.
                thisWS = self.wb.create_sheet(paramName)
                #thisWS.title = paramName
                thisWS['A1'] = "File"
                thisWS['B1'] = "Model"
                thisWS['C1'] = "Parameter"
                thisWS['D1'] = "Value"
                thisWS['E1'] = "Lower"
                thisWS['F1'] = "Upper"
                
                thisWS['A2'] = fileName
                thisWS['B2'] = modelName
                thisWS['C2'] = paramName
                thisWS['D2'] = str(paramValue)
                thisWS['E2'] = str(paramLower)
                thisWS['F2'] = str(paramUpper)
            else:
                #Worksheet already exists, 
                #so retrieve it
                thisWS = self.wb[paramName]
                #get next empty row
                row_count = thisWS.max_row
                nextRow = row_count + 1
                thisWS['A' + str(nextRow)] = fileName
                thisWS['B' + str(nextRow)] = modelName
                thisWS['C' + str(nextRow)] = paramName
                thisWS['D' + str(nextRow)] = str(paramValue)
                thisWS['E' + str(nextRow)] = str(paramLower)
                thisWS['F' + str(nextRow)] = str(paramUpper)
                        
            logger.info('In module ' + __name__ 
                    + '.RecordParameterValues when paramater = ' + paramName)
        except Exception as e:
            print('ExcelWriter.RecordParameterValues when paramater = ' 
                  + paramName + str(e)) 
            logger.error('ExcelWriter.RecordParameterValues when paramater = ' 
                         + paramName + str(e)) 

    def SaveSpreadSheet(self): 
        """ Saves the workbook as an Excel spreadsheet at fullFilePath"""
        try:
            self.wb.save(self.fullFilePath)
            logger.info('In module ' + __name__ 
                    + '. SaveSpreadSheet.')
        except Exception as e:
            print('ExcelWriter.SaveSpreadSheet: ' + str(e)) 
            logger.error('ExcelWriter.SaveSpreadSheet: ' + str(e)) 
            





